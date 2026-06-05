import base64
import json
import os
import csv
import io
from operator import add
from typing import Annotated, TypedDict
from datetime import datetime
import anthropic
import requests
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from langgraph.graph import END, StateGraph
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

LLM_MODEL = os.environ.get("LLM_MODEL", "anthropic-claude-sonnet-4-5-20250929_IrishConsolidation").strip('"')
TEMPERATURE = float(os.environ.get("TEMPERATURE", "0"))
MAX_TOKENS = int(os.environ.get("MAX_TOKENS_TO_SAMPLE", "18000"))
TENANT_KEY = os.environ["TENANT_KEY"].strip('"')
ASSET_ID = os.environ.get("ASSET_ID", "").strip('"')
USER_TYPE = os.environ.get("USER_TYPE", "machine").strip('"')

PROXY_BASE_URL = "https://cert-proxy-core.rag-us.use1.dev-searchplatform.nl.lexis.com/passthrough/anthropic"

client = anthropic.Anthropic(
    api_key=TENANT_KEY,
    base_url=PROXY_BASE_URL,
    default_headers={
        "x-asset-id": ASSET_ID,
        "x-user-type": USER_TYPE,
    },
)

ADO_ORG = os.environ["ADO_ORG"]
ADO_PROJECT = os.environ["ADO_PROJECT"]
ADO_PAT = os.environ["PAT"]
ADO_BASE = f"https://dev.azure.com/{ADO_ORG}/{ADO_PROJECT}/_apis/wit"
ADO_API_VERSION = os.environ.get("ADO_API_VERSION", "7.0").strip('"')
API_VER = f"api-version={ADO_API_VERSION}"

# Keep comments API configurable because Azure DevOps comments preview versions can vary by org.
COMMENTS_API_VER = os.environ.get(
    "ADO_COMMENTS_API_VERSION",
    "api-version=7.0-preview.3",
).strip('"')

# Keep only recent chat context so old success messages do not bias the model.
CHAT_HISTORY_WINDOW = int(os.environ.get("CHAT_HISTORY_WINDOW", "8"))

DEFAULT_ITERATION_PATH = os.environ.get("DEFAULT_ITERATION_PATH", "").strip('"')
DEFAULT_SPRINT = os.environ.get("DEFAULT_SPRINT", "").strip('"')
DEFAULT_AREA_PATH = os.environ.get("DEFAULT_AREA_PATH", "").strip('"')

_allowed_projects_raw = os.environ.get("ALLOWED_PROJECTS", "")
ALLOWED_PROJECTS = {p.strip() for p in _allowed_projects_raw.split(",") if p.strip()}

USER_STORY_TASK_STATES = (
    "Not Started",
    "Ready",
    "In Progress",
    "Done",
    "Closed",
    "Removed",
)
USER_STORY_TASK_STATE_MAP = {s.lower(): s for s in USER_STORY_TASK_STATES}


# ── ADO helpers ───────────────────────────────────────────────────────────────

def _ado_headers(patch: bool = False) -> dict:
    token = base64.b64encode(f":{ADO_PAT}".encode()).decode()
    headers = {"Authorization": f"Basic {token}"}
    if patch:
        headers["Content-Type"] = "application/json-patch+json"
    return headers


def _build_patch(field_map: dict) -> list:
    """Convert a field-value dict to ADO JSON Patch operations."""
    return [{"op": "add", "path": f"/fields/{k}", "value": v} for k, v in field_map.items()]


FIELD_NAMES = {
    "title":               "System.Title",
    "description":         "System.Description",
    "repro_steps":         "Microsoft.VSTS.TCM.ReproSteps",
    "acceptance_criteria": "Microsoft.VSTS.Common.AcceptanceCriteria",
    "story_points":        "Microsoft.VSTS.Scheduling.StoryPoints",
    "priority":            "Microsoft.VSTS.Common.Priority",
    "activity":            "Microsoft.VSTS.Common.Activity",
    "assignee":            "System.AssignedTo",
    "tags":                "System.Tags",
    "state":               "System.State",
    "reason":              "System.Reason",
    "area_path":           "System.AreaPath",
    "iteration_path":      "System.IterationPath",
    "remaining_work":      "Microsoft.VSTS.Scheduling.RemainingWork",
    "original_estimate":   "Microsoft.VSTS.Scheduling.OriginalEstimate",
    "completed_work":      "Microsoft.VSTS.Scheduling.CompletedWork",
    "business_value":      "Microsoft.VSTS.Common.BusinessValue",
    "effort":              "Microsoft.VSTS.Scheduling.Effort",
    "risk":                "Microsoft.VSTS.Common.Risk",
    "severity":            "Microsoft.VSTS.Common.Severity",
    "start_date":          "Microsoft.VSTS.Scheduling.StartDate",
    "target_date":         "Microsoft.VSTS.Scheduling.TargetDate",
}


def _work_item_ref_url(work_item_id: int) -> str:
    return f"https://dev.azure.com/{ADO_ORG}/{ADO_PROJECT}/_apis/wit/workitems/{work_item_id}"


def _work_item_edit_url(work_item_id: int) -> str:
    return f"https://dev.azure.com/{ADO_ORG}/{ADO_PROJECT}/_workitems/edit/{work_item_id}"


def _get_work_item(work_item_id: int, expand: str = "all") -> dict:
    res = requests.get(
        f"{ADO_BASE}/workitems/{work_item_id}?{API_VER}&$expand={expand}",
        headers=_ado_headers(),
        timeout=30,
    )
    res.raise_for_status()
    return res.json()


def _to_assignee(value):
    if isinstance(value, dict):
        return value.get("displayName") or value.get("uniqueName")
    return value


def _summarize_item(data: dict) -> dict:
    fields = data.get("fields", {})
    return {
        "id": data.get("id"),
        "type": fields.get("System.WorkItemType"),
        "title": fields.get("System.Title"),
        "state": fields.get("System.State"),
        "priority": fields.get("Microsoft.VSTS.Common.Priority"),
        "assignee": _to_assignee(fields.get("System.AssignedTo")),
        "tags": fields.get("System.Tags"),
        "url": data.get("_links", {}).get("html", {}).get("href") or _work_item_edit_url(data.get("id")),
    }


def _normalize_user_story_task_state(value: str) -> str | None:
    if value is None:
        return None
    return USER_STORY_TASK_STATE_MAP.get(str(value).strip().lower())


def _normalize_assignee_value(value):
    if value is None:
        return None
    if isinstance(value, dict):
        value = value.get("uniqueName") or value.get("displayName") or value.get("name")
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    return value


def _normalize_user_story_inputs(inputs: dict) -> dict:
    payload = dict(inputs)

    # Accept common alias keys from natural language tool calls.
    assignee_aliases = (
        "assigned_to",
        "assignedTo",
        "assignedto",
        "owner",
    )
    if payload.get("assignee") is None:
        for alias in assignee_aliases:
            if payload.get(alias) is not None:
                payload["assignee"] = payload.get(alias)
                break

    payload["assignee"] = _normalize_assignee_value(payload.get("assignee"))
    return payload


def _default_iteration_path() -> str | None:
    if not DEFAULT_ITERATION_PATH:
        return None
    if not DEFAULT_SPRINT:
        return DEFAULT_ITERATION_PATH

    suffix = f"\\{DEFAULT_SPRINT}"
    if DEFAULT_ITERATION_PATH.endswith(suffix):
        return DEFAULT_ITERATION_PATH
    return f"{DEFAULT_ITERATION_PATH}{suffix}"


def _default_area_path(project_name: str | None) -> str | None:
    if not DEFAULT_AREA_PATH:
        return None

    if "{project_name}" in DEFAULT_AREA_PATH:
        if not project_name:
            return None
        return DEFAULT_AREA_PATH.format(project_name=project_name)

    if project_name:
        return f"{DEFAULT_AREA_PATH}\\{project_name}"
    return DEFAULT_AREA_PATH


def _update_work_item(work_item_id: int, patch: list) -> dict:
    res = requests.patch(
        f"{ADO_BASE}/workitems/{work_item_id}?{API_VER}",
        headers=_ado_headers(patch=True),
        json=patch,
        timeout=30,
    )
    res.raise_for_status()
    return res.json()


def _delete_work_item(inputs: dict, expected_type: str | None = None) -> str:
    work_item_id = int(inputs["id"])
    action = inputs.get("action", "soft-delete")

    if action == "soft-delete":
        if expected_type:
            try:
                existing = _get_work_item(work_item_id)
            except requests.HTTPError:
                return json.dumps({
                    "error": f"Work item {work_item_id} not found or already deleted.",
                })
            actual_type = existing.get("fields", {}).get("System.WorkItemType")
            if actual_type != expected_type:
                return json.dumps({
                    "error": f"Work item {work_item_id} is type '{actual_type}', expected '{expected_type}'.",
                })

        res = requests.delete(
            f"{ADO_BASE}/workitems/{work_item_id}?{API_VER}",
            headers=_ado_headers(),
            timeout=30,
        )
        res.raise_for_status()
        return json.dumps({
            "id": work_item_id,
            "action": action,
            "status": "✅ deleted",
        })

    if action == "restore":
        res = requests.patch(
            f"{ADO_BASE}/recyclebin/{work_item_id}?{API_VER}",
            headers={**_ado_headers(), "Content-Type": "application/json"},
            json={"isDeleted": False},
            timeout=30,
        )
        res.raise_for_status()
        data = res.json()
        return json.dumps({
            "id": data.get("id", work_item_id),
            "action": action,
            "status": "✅ restored",
            "url": _work_item_edit_url(data.get("id", work_item_id)),
        })

    if action == "permanent-delete":
        res = requests.delete(
            f"{ADO_BASE}/workitems/{work_item_id}?destroy=true&{API_VER}",
            headers=_ado_headers(),
            timeout=30,
        )
        res.raise_for_status()
        return json.dumps({
            "id": work_item_id,
            "action": action,
            "status": "✅ permanently deleted",
        })

    return json.dumps({"error": "Invalid action. Use soft-delete, restore, or permanent-delete."})


# ── Tool implementations ──────────────────────────────────────────────────────

def ado_create_user_story(inputs: dict) -> str:
    inputs = _normalize_user_story_inputs(inputs)
    feature_id = inputs.get("feature_id")
    if feature_id is None:
        for key in ("parent_feature_id", "featureId", "parentId"):
            if inputs.get(key) is not None:
                feature_id = inputs.get(key)
                inputs["feature_id"] = feature_id
                break
    project_name = inputs.get("project_name")

    if feature_id is None and not project_name:
        return json.dumps({
            "error": "Either feature_id or project_name is required for create_user_story.",
        })

    if project_name and ALLOWED_PROJECTS and project_name not in ALLOWED_PROJECTS:
        return json.dumps({
            "error": f"Invalid project_name '{project_name}'. Allowed values: {', '.join(sorted(ALLOWED_PROJECTS))}",
        })

    # Enforce defaults unless the caller explicitly provides values.
    if inputs.get("story_points") is None:
        inputs["story_points"] = 3
    if inputs.get("priority") is None:
        inputs["priority"] = 3
    if inputs.get("iteration_path") is None:
        iteration_path = _default_iteration_path()
        if iteration_path:
            inputs["iteration_path"] = iteration_path

    if inputs.get("area_path") is None:
        area_path = _default_area_path(project_name)
        if area_path:
            inputs["area_path"] = area_path
        else:
            return json.dumps({
                "error": "area_path is required, or configure DEFAULT_AREA_PATH (optionally with {project_name}) and provide project_name.",
            })

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in inputs.items()
        if k in FIELD_NAMES and v is not None
    })

    # Link to parent feature when provided.
    if feature_id is not None:
        patch.append({
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "System.LinkTypes.Hierarchy-Reverse",
                "url": _work_item_ref_url(int(feature_id)),
            },
        })

    res = requests.post(
        f"{ADO_BASE}/workitems/$User%20Story?{API_VER}",
        headers=_ado_headers(patch=True),
        json=patch,
        timeout=30,
    )
    res.raise_for_status()
    data = res.json()
    return json.dumps({
        "id": data["id"],
        "title": data["fields"]["System.Title"],
        "url": data["_links"]["html"]["href"],
        "status": "✅ created",
    })


def ado_update_user_story(inputs: dict) -> str:
    payload = dict(inputs)

    story_id = payload.get("story_id")
    if story_id is None:
        for key in ("user_story_id", "userStoryId", "id"):
            if payload.get(key) is not None:
                story_id = payload.get(key)
                break

    if story_id is None:
        return json.dumps({"error": "story_id is required for update_user_story."})

    story_id = int(story_id)
    comment_text = payload.pop("comment", None)

    payload = _normalize_user_story_inputs(payload)

    if payload.get("state") is not None:
        normalized_state = _normalize_user_story_task_state(payload["state"])
        if normalized_state is None:
            return json.dumps({
                "error": "Invalid state for User Story. Allowed values: " + ", ".join(USER_STORY_TASK_STATES),
            })
        payload["state"] = normalized_state

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in payload.items()
        if k in FIELD_NAMES and v is not None and k != "story_id"
    })

    result: dict = {}

    if patch:
        data = _update_work_item(story_id, patch)
        result = {
            "id": data["id"],
            "title": data["fields"].get("System.Title"),
            "state": data["fields"].get("System.State"),
            "url": data.get("_links", {}).get("html", {}).get("href") or _work_item_edit_url(story_id),
            "status": "✅ updated",
        }
    else:
        result = {"id": story_id, "status": "no field changes"}

    if comment_text:
        comment_result = _post_work_item_comment(story_id, comment_text)
        result["comment"] = comment_result

        if comment_result.get("error"):
            result["error"] = comment_result["error"]

    if not patch and not comment_text:
        return json.dumps({"error": "No fields or comment provided to update."})

    return json.dumps(result)

def _post_work_item_comment(work_item_id: int, text: str) -> dict:
    """Post a comment to a work item's discussion using the Comments API."""
    res = requests.post(
        f"{ADO_BASE}/workitems/{work_item_id}/comments?{COMMENTS_API_VER}",
        headers={**_ado_headers(), "Content-Type": "application/json"},
        json={"text": text},
        timeout=30,
    )

    if not res.ok:
        return {
            "work_item_id": work_item_id,
            "error": f"Comment post failed: {res.status_code} {res.text}",
        }

    c = res.json()
    return {
        "work_item_id": work_item_id,
        "comment_id": c.get("id"),
        "created": c.get("createdDate"),
        "status": "✅ comment posted",
    }

def ado_add_comment(inputs: dict) -> str:
    """Standalone tool: post a comment to any work item."""
    work_item_id = int(inputs["work_item_id"])
    text = inputs["text"]
    result = _post_work_item_comment(work_item_id, text)
    return json.dumps(result)


def ado_get_user_story(inputs: dict) -> str:
    story_id = inputs.get("story_id")
    if story_id is None:
        for key in ("user_story_id", "userStoryId", "id"):
            if inputs.get(key) is not None:
                story_id = inputs.get(key)
                break
    if story_id is None:
        return json.dumps({"error": "story_id is required for get_user_story."})
    story_id = int(story_id)
    data = _get_work_item(story_id)
    f = data["fields"]
    return json.dumps({
        "id": story_id,
        "title": f.get("System.Title"),
        "state": f.get("System.State"),
        "reason": f.get("System.Reason"),
        "priority": f.get("Microsoft.VSTS.Common.Priority"),
        "story_points": f.get("Microsoft.VSTS.Scheduling.StoryPoints"),
        "assignee": f.get("System.AssignedTo", {}).get("displayName") if isinstance(f.get("System.AssignedTo"), dict) else f.get("System.AssignedTo"),
        "tags": f.get("System.Tags"),
        "description": f.get("System.Description"),
        "acceptance_criteria": f.get("Microsoft.VSTS.Common.AcceptanceCriteria"),
        "area_path": f.get("System.AreaPath"),
        "iteration_path": f.get("System.IterationPath"),
        "url": _work_item_edit_url(story_id),
    })


def ado_delete_user_story(inputs: dict) -> str:
    story_id = inputs.get("story_id")
    if story_id is None:
        for key in ("user_story_id", "userStoryId", "id"):
            if inputs.get(key) is not None:
                story_id = inputs.get(key)
                break
    if story_id is None:
        return json.dumps({"error": "story_id is required for delete_user_story."})

    mapped = {
        "id": story_id,
        "action": inputs.get("action", "soft-delete"),
    }
    return _delete_work_item(mapped, expected_type="User Story")


def ado_create_bug(inputs: dict) -> str:
    if not inputs.get("title"):
        return json.dumps({"error": "title is required for create_bug."})

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in inputs.items()
        if k in FIELD_NAMES and v is not None
    })

    res = requests.post(
        f"{ADO_BASE}/workitems/$Bug?{API_VER}",
        headers=_ado_headers(patch=True),
        json=patch,
        timeout=30,
    )
    res.raise_for_status()
    data = res.json()
    out = _summarize_item(data)
    out["status"] = "✅ created"
    return json.dumps(out)


def ado_get_bug(inputs: dict) -> str:
    data = _get_work_item(int(inputs["bug_id"]))
    out = _summarize_item(data)
    f = data.get("fields", {})
    out["description"]    = f.get("System.Description")
    out["repro_steps"]    = f.get("Microsoft.VSTS.TCM.ReproSteps")
    out["severity"]       = f.get("Microsoft.VSTS.Common.Severity")
    out["area_path"]      = f.get("System.AreaPath")
    out["iteration_path"] = f.get("System.IterationPath")
    return json.dumps(out)


def ado_update_bug(inputs: dict) -> str:
    bug_id = int(inputs["bug_id"])
    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in inputs.items()
        if k in FIELD_NAMES and v is not None and k != "bug_id"
    })
    if not patch:
        return json.dumps({"error": "No fields provided to update."})

    data = _update_work_item(bug_id, patch)
    out = _summarize_item(data)
    out["status"] = "✅ updated"
    return json.dumps(out)


def ado_delete_bug(inputs: dict) -> str:
    mapped = {
        "id": inputs["bug_id"],
        "action": inputs.get("action", "soft-delete"),
    }
    return _delete_work_item(mapped, expected_type="Bug")


def ado_create_epic(inputs: dict) -> str:
    payload = dict(inputs)
    if payload.get("title") is None:
        for key in ("epic_title", "name"):
            if payload.get(key) is not None:
                payload["title"] = payload.get(key)
                break

    if payload.get("title") is None:
        return json.dumps({"error": "title is required for create_epic."})

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in payload.items()
        if k in FIELD_NAMES and v is not None
    })

    res = requests.post(
        f"{ADO_BASE}/workitems/$Epic?{API_VER}",
        headers=_ado_headers(patch=True),
        json=patch,
        timeout=30,
    )
    res.raise_for_status()
    data = res.json()
    out = _summarize_item(data)
    out["status"] = "✅ created"
    return json.dumps(out)


def ado_get_epic(inputs: dict) -> str:
    epic_id = int(inputs["epic_id"])
    data = _get_work_item(epic_id)
    out = _summarize_item(data)
    f = data.get("fields", {})
    out["description"] = f.get("System.Description")
    out["acceptance_criteria"] = f.get("Microsoft.VSTS.Common.AcceptanceCriteria")
    out["area_path"] = f.get("System.AreaPath")
    out["iteration_path"] = f.get("System.IterationPath")
    out["start_date"] = f.get("Microsoft.VSTS.Scheduling.StartDate")
    out["target_date"] = f.get("Microsoft.VSTS.Scheduling.TargetDate")
    out["business_value"] = f.get("Microsoft.VSTS.Common.BusinessValue")
    out["effort"] = f.get("Microsoft.VSTS.Scheduling.Effort")
    out["risk"] = f.get("Microsoft.VSTS.Common.Risk")
    return json.dumps(out)


def ado_update_epic(inputs: dict) -> str:
    epic_id = int(inputs["epic_id"])
    payload = dict(inputs)

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in payload.items()
        if k in FIELD_NAMES and v is not None and k != "epic_id"
    })

    if not patch:
        return json.dumps({"error": "No fields provided to update."})

    data = _update_work_item(epic_id, patch)
    out = _summarize_item(data)
    out["status"] = "✅ updated"
    return json.dumps(out)


def ado_delete_epic(inputs: dict) -> str:
    mapped = {
        "id": inputs["epic_id"],
        "action": inputs.get("action", "soft-delete"),
    }
    return _delete_work_item(mapped, expected_type="Epic")


def _find_parent_relation_index(relations: list[dict]) -> int | None:
    for idx, rel in enumerate(relations or []):
        if rel.get("rel") == "System.LinkTypes.Hierarchy-Reverse":
            return idx
    return None


def ado_create_feature(inputs: dict) -> str:
    payload = dict(inputs)
    if payload.get("title") is None:
        for key in ("feature_title", "name"):
            if payload.get(key) is not None:
                payload["title"] = payload.get(key)
                break

    if payload.get("title") is None:
        return json.dumps({"error": "title is required for create_feature."})

    parent_epic_id = payload.get("parent_epic_id")
    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in payload.items()
        if k in FIELD_NAMES and v is not None
    })

    if parent_epic_id is not None:
        parent = _get_work_item(int(parent_epic_id), expand="none")
        parent_type = parent.get("fields", {}).get("System.WorkItemType")
        if parent_type != "Epic":
            return json.dumps({
                "error": f"parent_epic_id {int(parent_epic_id)} is type '{parent_type}', expected 'Epic'.",
            })
        patch.append({
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "System.LinkTypes.Hierarchy-Reverse",
                "url": _work_item_ref_url(int(parent_epic_id)),
            },
        })

    res = requests.post(
        f"{ADO_BASE}/workitems/$Feature?{API_VER}",
        headers=_ado_headers(patch=True),
        json=patch,
        timeout=30,
    )
    res.raise_for_status()
    data = res.json()
    out = _summarize_item(data)
    out["status"] = "✅ created"
    return json.dumps(out)


def ado_get_feature(inputs: dict) -> str:
    feature_id = int(inputs["feature_id"])
    data = _get_work_item(feature_id)
    out = _summarize_item(data)
    f = data.get("fields", {})
    out["description"] = f.get("System.Description")
    out["acceptance_criteria"] = f.get("Microsoft.VSTS.Common.AcceptanceCriteria")
    out["area_path"] = f.get("System.AreaPath")
    out["iteration_path"] = f.get("System.IterationPath")
    out["start_date"] = f.get("Microsoft.VSTS.Scheduling.StartDate")
    out["target_date"] = f.get("Microsoft.VSTS.Scheduling.TargetDate")
    out["business_value"] = f.get("Microsoft.VSTS.Common.BusinessValue")
    out["effort"] = f.get("Microsoft.VSTS.Scheduling.Effort")
    out["risk"] = f.get("Microsoft.VSTS.Common.Risk")

    parent_id = None
    child_story_ids = []
    _CHILD_RELS = {"System.LinkTypes.Hierarchy-Forward", "Child"}
    _PARENT_RELS = {"System.LinkTypes.Hierarchy-Reverse", "Parent"}
    for rel in data.get("relations", []):
        rel_type = rel.get("rel", "")
        url = rel.get("url", "")
        if rel_type in _PARENT_RELS:
            if "/workitems/" in url.lower():
                try:
                    parent_id = int(url.rsplit("/", 1)[-1])
                except ValueError:
                    parent_id = None
        elif rel_type in _CHILD_RELS:
            if "/workitems/" in url.lower():
                try:
                    child_story_ids.append(int(url.rsplit("/", 1)[-1]))
                except ValueError:
                    pass
    out["parent_epic_id"] = parent_id
    out["child_story_ids"] = child_story_ids
    out["child_story_count"] = len(child_story_ids)
    return json.dumps(out)


# ADO may return the short form "Child" / "Parent" or the full
# "System.LinkTypes.Hierarchy-Forward" / "Hierarchy-Reverse" depending on
# the API version and expansion level.
_CHILD_REL_TYPES = {"System.LinkTypes.Hierarchy-Forward", "Child"}
_PARENT_REL_TYPES = {"System.LinkTypes.Hierarchy-Reverse", "Parent"}

def _normalize_report_item_types(value) -> set[str]:
    if not value:
        return set()

    if isinstance(value, str):
        raw_items = [v.strip() for v in value.split(",")]
    elif isinstance(value, list):
        raw_items = [str(v).strip() for v in value]
    else:
        return set()

    normalized = set()

    for item in raw_items:
        lowered = item.lower()

        if lowered in {"task", "tasks"}:
            normalized.add("Task")
        elif lowered in {"bug", "bugs"}:
            normalized.add("Bug")
        elif lowered in {"story", "user story", "user stories"}:
            normalized.add("User Story")
        elif lowered in {"feature", "features"}:
            normalized.add("Feature")
        elif lowered in {"epic", "epics"}:
            normalized.add("Epic")

    return normalized


def _bool_value(value, default=False) -> bool:
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "1", "y"}

    return bool(value)


def _get_child_ids_from_work_item(data: dict) -> list[int]:
    child_ids = []

    for rel in data.get("relations", []) or []:
        if rel.get("rel") in _CHILD_REL_TYPES:
            url = rel.get("url", "")

            if "/workitems/" in url.lower():
                try:
                    child_ids.append(int(url.rsplit("/", 1)[-1]))
                except ValueError:
                    pass

    return child_ids


def _work_item_report_row(data: dict, parent_id: int | None = None, level: int = 0) -> dict:
    fields = data.get("fields", {})
    work_item_id = data.get("id")

    return {
        "level": level,
        "parent_id": parent_id,
        "id": work_item_id,
        "type": fields.get("System.WorkItemType"),
        "title": fields.get("System.Title"),
        "state": fields.get("System.State"),
        "reason": fields.get("System.Reason"),
        "priority": fields.get("Microsoft.VSTS.Common.Priority"),
        "assignee": _to_assignee(fields.get("System.AssignedTo")),
        "tags": fields.get("System.Tags"),

        "story_points": fields.get("Microsoft.VSTS.Scheduling.StoryPoints"),
        "business_value": fields.get("Microsoft.VSTS.Common.BusinessValue"),
        "effort": fields.get("Microsoft.VSTS.Scheduling.Effort"),
        "risk": fields.get("Microsoft.VSTS.Common.Risk"),

        "activity": fields.get("Microsoft.VSTS.Common.Activity"),
        "original_estimate": fields.get("Microsoft.VSTS.Scheduling.OriginalEstimate"),
        "completed_work": fields.get("Microsoft.VSTS.Scheduling.CompletedWork"),
        "remaining_work": fields.get("Microsoft.VSTS.Scheduling.RemainingWork"),

        "severity": fields.get("Microsoft.VSTS.Common.Severity"),
        "start_date": fields.get("Microsoft.VSTS.Scheduling.StartDate"),
        "target_date": fields.get("Microsoft.VSTS.Scheduling.TargetDate"),

        "area_path": fields.get("System.AreaPath"),
        "iteration_path": fields.get("System.IterationPath"),

        "url": _work_item_edit_url(work_item_id),
    }

def ado_get_work_item_relations(inputs: dict) -> str:
    """Debug tool: return raw relation types on a work item."""
    work_item_id = int(inputs["work_item_id"])
    data = _get_work_item(work_item_id)
    relations = [
        {"rel": r.get("rel"), "url": r.get("url", "").rsplit("/", 1)[-1]}
        for r in (data.get("relations") or [])
    ]
    return json.dumps({"work_item_id": work_item_id, "relations": relations, "count": len(relations)})

def ado_get_work_item_children(inputs: dict) -> str:
    """
    Generic child/hierarchy report for any Azure DevOps work item.
    Works for Epic, Feature, User Story, Bug, Task, or any work item with child links.
    """

    work_item_id = inputs.get("work_item_id")

    if work_item_id is None:
        for key in (
            "id",
            "parent_id",
            "epic_id",
            "feature_id",
            "story_id",
            "user_story_id",
            "task_id",
            "bug_id",
        ):
            if inputs.get(key) is not None:
                work_item_id = inputs.get(key)
                break

    if work_item_id is None:
        return json.dumps({
            "error": "work_item_id is required for get_work_item_children."
        })

    work_item_id = int(work_item_id)

    recursive = _bool_value(inputs.get("recursive"), default=False)
    include_parent = _bool_value(inputs.get("include_parent"), default=False)
    item_types = _normalize_report_item_types(inputs.get("item_types"))

    visited: set[int] = set()

    def collect_children(parent_id: int, level: int) -> list[dict]:
        if parent_id in visited:
            return []

        visited.add(parent_id)

        try:
            parent_data = _get_work_item(parent_id, expand="all")
        except requests.HTTPError as e:
            return [{
                "level": level,
                "parent_id": None,
                "id": parent_id,
                "error": f"Could not fetch parent: {e.response.status_code} {e.response.text}",
            }]

        child_ids = _get_child_ids_from_work_item(parent_data)
        rows = []

        for child_id in child_ids:
            try:
                child_data = _get_work_item(child_id, expand="all")

                child_row = _work_item_report_row(
                    child_data,
                    parent_id=parent_id,
                    level=level + 1,
                )

                child_type = child_row.get("type")

                if not item_types or child_type in item_types:
                    rows.append(child_row)

                if recursive:
                    rows.extend(collect_children(child_id, level + 1))

            except requests.HTTPError as e:
                rows.append({
                    "level": level + 1,
                    "parent_id": parent_id,
                    "id": child_id,
                    "error": f"Could not fetch child: {e.response.status_code} {e.response.text}",
                })

        return rows

    parent_data = _get_work_item(work_item_id, expand="all")
    parent_row = _work_item_report_row(parent_data, parent_id=None, level=0)

    child_rows = collect_children(work_item_id, level=0)

    rows = []

    if include_parent:
        rows.append(parent_row)

    rows.extend(child_rows)

    return json.dumps({
        "work_item_id": work_item_id,
        "parent": parent_row,
        "recursive": recursive,
        "include_parent": include_parent,
        "item_types": sorted(item_types),
        "rows": rows,
        "children": child_rows,
        "count": len(child_rows),
    })

def ado_get_feature_children(inputs: dict) -> str:
    feature_id = int(inputs["feature_id"])
    data = _get_work_item(feature_id)

    child_ids = []
    for rel in data.get("relations", []):
        if rel.get("rel") in _CHILD_REL_TYPES:
            url = rel.get("url", "")
            if "/workitems/" in url.lower():
                try:
                    child_ids.append(int(url.rsplit("/", 1)[-1]))
                except ValueError:
                    pass

    if not child_ids:
        return json.dumps({
            "feature_id": feature_id,
            "children": [],
            "count": 0,
        })

    children = []
    for child_id in child_ids:
        try:
            child_data = _get_work_item(child_id, expand="none")
            f = child_data.get("fields", {})
            children.append({
                "id": child_data.get("id"),
                "type": f.get("System.WorkItemType"),
                "title": f.get("System.Title"),
                "state": f.get("System.State"),
                "priority": f.get("Microsoft.VSTS.Common.Priority"),
                "story_points": f.get("Microsoft.VSTS.Scheduling.StoryPoints"),
                "assignee": _to_assignee(f.get("System.AssignedTo")),
                "tags": f.get("System.Tags"),
                "url": _work_item_edit_url(child_data.get("id")),
            })
        except requests.HTTPError:
            children.append({"id": child_id, "error": "could not fetch"})

    return json.dumps({
        "feature_id": feature_id,
        "children": children,
        "count": len(children),
    })


def ado_update_feature(inputs: dict) -> str:
    feature_id = int(inputs["feature_id"])
    payload = dict(inputs)
    new_parent_epic_id = payload.get("parent_epic_id")

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in payload.items()
        if k in FIELD_NAMES and v is not None and k != "feature_id"
    })

    if new_parent_epic_id is not None:
        parent = _get_work_item(int(new_parent_epic_id), expand="none")
        parent_type = parent.get("fields", {}).get("System.WorkItemType")
        if parent_type != "Epic":
            return json.dumps({
                "error": f"parent_epic_id {int(new_parent_epic_id)} is type '{parent_type}', expected 'Epic'.",
            })

        feature = _get_work_item(feature_id)
        parent_idx = _find_parent_relation_index(feature.get("relations", []))
        if parent_idx is not None:
            patch.append({"op": "remove", "path": f"/relations/{parent_idx}"})
        patch.append({
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "System.LinkTypes.Hierarchy-Reverse",
                "url": _work_item_ref_url(int(new_parent_epic_id)),
            },
        })

    if not patch:
        return json.dumps({"error": "No fields provided to update."})

    data = _update_work_item(feature_id, patch)
    out = _summarize_item(data)
    out["status"] = "✅ updated"
    return json.dumps(out)


def ado_delete_feature(inputs: dict) -> str:
    mapped = {
        "id": inputs["feature_id"],
        "action": inputs.get("action", "soft-delete"),
    }
    return _delete_work_item(mapped, expected_type="Feature")


def ado_create_task(inputs: dict) -> str:
    payload = dict(inputs)

    # Normalize common key variants used by natural-language tool calls.
    if payload.get("title") is None:
        for key in ("task_title", "name"):
            if payload.get(key) is not None:
                payload["title"] = payload.get(key)
                break

    if payload.get("parent_id") is None:
        for key in ("parentId", "story_id", "user_story_id"):
            if payload.get(key) is not None:
                payload["parent_id"] = payload.get(key)
                break

    if payload.get("remaining_work") is None and payload.get("remainingWork") is not None:
        payload["remaining_work"] = payload.get("remainingWork")

    if isinstance(payload.get("tags"), list):
        payload["tags"] = "; ".join(str(t).strip() for t in payload["tags"] if str(t).strip()) or None

    if isinstance(payload.get("remaining_work"), str):
        text = payload["remaining_work"].strip()
        if text:
            try:
                payload["remaining_work"] = float(text) if "." in text else int(text)
            except ValueError:
                pass

    if payload.get("title") is None:
        return json.dumps({"error": "title is required for create_task."})

    inherited_fields = []

    if payload.get("parent_id") is not None:
        parent = _get_work_item(int(payload["parent_id"]), expand="none")
        parent_fields = parent.get("fields", {})
        if parent_fields.get("System.WorkItemType") == "User Story":
            if payload.get("area_path") is None and parent_fields.get("System.AreaPath"):
                payload["area_path"] = parent_fields.get("System.AreaPath")
                inherited_fields.append("area_path")
            if payload.get("iteration_path") is None and parent_fields.get("System.IterationPath"):
                payload["iteration_path"] = parent_fields.get("System.IterationPath")
                inherited_fields.append("iteration_path")

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in payload.items()
        if k in FIELD_NAMES and v is not None
    })

    if payload.get("parent_id") is not None:
        patch.append({
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "System.LinkTypes.Hierarchy-Reverse",
                "url": _work_item_ref_url(int(payload["parent_id"])),
            },
        })

    res = requests.post(
        f"{ADO_BASE}/workitems/$Task?{API_VER}",
        headers=_ado_headers(patch=True),
        json=patch,
        timeout=30,
    )
    res.raise_for_status()
    data = res.json()
    out = _summarize_item(data)
    if inherited_fields:
        out["inherited_from_parent_user_story"] = inherited_fields
    out["status"] = "✅ created"
    return json.dumps(out)


def ado_get_task(inputs: dict) -> str:
    task_id = int(inputs["task_id"])

    data = _get_work_item(task_id)
    f = data.get("fields", {})

    out = _summarize_item(data)

    # Standard task fields
    out["reason"] = f.get("System.Reason")
    out["activity"] = f.get("Microsoft.VSTS.Common.Activity")
    out["description"] = f.get("System.Description")

    # Work tracking fields
    out["original_estimate"] = f.get(
        "Microsoft.VSTS.Scheduling.OriginalEstimate"
    )
    out["remaining_work"] = f.get(
        "Microsoft.VSTS.Scheduling.RemainingWork"
    )
    out["completed_work"] = f.get(
        "Microsoft.VSTS.Scheduling.CompletedWork"
    )

    # Classification fields
    out["area_path"] = f.get("System.AreaPath")
    out["iteration_path"] = f.get("System.IterationPath")

    # Parent relationship
    parent_id = None
    parent_type = None

    for rel in data.get("relations", []) or []:
        if rel.get("rel") in {
            "System.LinkTypes.Hierarchy-Reverse",
            "Parent",
        }:
            url = rel.get("url", "")

            if "/workitems/" in url.lower():
                try:
                    parent_id = int(url.rsplit("/", 1)[-1])

                    try:
                        parent = _get_work_item(parent_id, expand="none")
                        parent_type = (
                            parent.get("fields", {})
                            .get("System.WorkItemType")
                        )
                    except Exception:
                        pass

                except ValueError:
                    pass

            break

    out["parent_id"] = parent_id
    out["parent_type"] = parent_type

    return json.dumps(out)

def ado_update_task(inputs: dict) -> str:
    payload = dict(inputs)

    task_id = int(payload["task_id"])
    comment_text = payload.pop("comment", None)

    if payload.get("state") is not None:
        normalized_state = _normalize_user_story_task_state(payload["state"])
        if normalized_state is None:
            return json.dumps({
                "error": "Invalid state for Task. Allowed values: " + ", ".join(USER_STORY_TASK_STATES),
            })
        payload["state"] = normalized_state

    for num_field in ("remaining_work", "original_estimate", "completed_work"):
        if isinstance(payload.get(num_field), str):
            text = payload[num_field].strip()
            try:
                payload[num_field] = float(text) if "." in text else int(text)
            except ValueError:
                payload.pop(num_field, None)

    patch = _build_patch({
        FIELD_NAMES[k]: v
        for k, v in payload.items()
        if k in FIELD_NAMES and v is not None and k != "task_id"
    })

    if not patch and not comment_text:
        return json.dumps({"error": "No fields or comment provided to update."})

    result: dict = {}

    if patch:
        data = _update_work_item(task_id, patch)
        result = _summarize_item(data)
        result["status"] = "✅ updated"
    else:
        result = {"id": task_id, "status": "no field changes"}

    if comment_text:
        comment_result = _post_work_item_comment(task_id, comment_text)
        result["comment"] = comment_result

        if comment_result.get("error"):
            result["error"] = comment_result["error"]

    return json.dumps(result)

def ado_delete_task(inputs: dict) -> str:
    mapped = {
        "id": inputs["task_id"],
        "action": inputs.get("action", "soft-delete"),
    }
    return _delete_work_item(mapped, expected_type="Task")


LINK_TYPES = {
    "parent": "System.LinkTypes.Hierarchy-Reverse",
    "dependency": "System.LinkTypes.Dependency-Reverse",
    "related": "System.LinkTypes.Related",
}


ALLOWED_PARENT_LINKS = {
    "Feature": {"Epic"},
    "User Story": {"Feature"},
    "Task": {"User Story"},
    "Bug": {"Feature", "User Story"},
}


def _normalize_linking_inputs(inputs: dict) -> dict:
    payload = dict(inputs)

    # Support both camelCase and snake_case keys.
    if payload.get("sourceId") is None and payload.get("source_id") is not None:
        payload["sourceId"] = payload["source_id"]
    if payload.get("targetId") is None and payload.get("target_id") is not None:
        payload["targetId"] = payload["target_id"]
    if payload.get("linkType") is None and payload.get("link_type") is not None:
        payload["linkType"] = payload["link_type"]

    # Backward compatibility with previous hierarchy tool payload shape.
    if payload.get("sourceId") is None and payload.get("child_id") is not None:
        payload["sourceId"] = payload["child_id"]
    if payload.get("targetId") is None and payload.get("parent_id") is not None:
        payload["targetId"] = payload["parent_id"]
    if payload.get("action") is None and payload.get("operation") is not None:
        payload["action"] = payload["operation"]

    if payload.get("linkType") is None:
        payload["linkType"] = "parent"

    if payload.get("comment") is None:
        payload["comment"] = "Linked by ADO Agent"

    return payload


def _validate_parent_link(source_id: int, target_id: int) -> dict | None:
    source = _get_work_item(source_id, expand="none")
    target = _get_work_item(target_id, expand="none")
    source_type = source.get("fields", {}).get("System.WorkItemType")
    target_type = target.get("fields", {}).get("System.WorkItemType")

    allowed_targets = ALLOWED_PARENT_LINKS.get(source_type, set())
    if target_type not in allowed_targets:
        return {
            "error": (
                f"Invalid parent link: {source_type} cannot link to {target_type}. "
                "Allowed parent links: Feature->Epic, User Story->Feature, Task->User Story, Bug->Feature/User Story."
            )
        }
    return None


def _find_relation_index(item: dict, relation: str, target_id: int) -> int:
    target_url = _work_item_ref_url(target_id).lower()
    for idx, rel in enumerate(item.get("relations", []) or []):
        if rel.get("rel") == relation and str(rel.get("url", "")).lower() == target_url:
            return idx
    return -1


def _link_work_items(source_id: int, target_id: int, link_type: str, comment: str) -> dict:
    relation = LINK_TYPES.get(link_type)
    if not relation:
        return {"error": f"Unsupported link type: {link_type}"}

    if link_type == "parent":
        validation_error = _validate_parent_link(source_id, target_id)
        if validation_error:
            return validation_error

    patch = [{
        "op": "add",
        "path": "/relations/-",
        "value": {
            "rel": relation,
            "url": _work_item_ref_url(target_id),
            "attributes": {"comment": comment},
        },
    }]
    return _update_work_item(source_id, patch)


def _unlink_work_items(source_id: int, target_id: int, link_type: str) -> dict:
    relation = LINK_TYPES.get(link_type)
    if not relation:
        return {"error": f"Unsupported link type: {link_type}"}

    item = _get_work_item(source_id)
    relation_index = _find_relation_index(item, relation, target_id)
    if relation_index == -1:
        return {"error": f"No {link_type} link found between #{source_id} and #{target_id}"}

    patch = [{"op": "remove", "path": f"/relations/{relation_index}"}]
    return _update_work_item(source_id, patch)


def ado_link_feature_hierarchy(inputs: dict) -> str:
    payload = _normalize_linking_inputs(inputs)

    if payload.get("sourceId") is None or payload.get("targetId") is None:
        return json.dumps({"error": "sourceId and targetId are required."})

    action = str(payload.get("action", "link")).strip().lower()
    if action not in {"link", "unlink"}:
        return json.dumps({"error": "Invalid action. Use link or unlink."})

    source_id = int(payload["sourceId"])
    target_id = int(payload["targetId"])
    link_type = str(payload.get("linkType", "parent")).strip().lower()

    if action == "link":
        data = _link_work_items(source_id, target_id, link_type, str(payload.get("comment", "Linked by ADO Agent")))
    else:
        data = _unlink_work_items(source_id, target_id, link_type)

    if data.get("error"):
        return json.dumps(data)

    return json.dumps({
        "action": action,
        "source_id": source_id,
        "target_id": target_id,
        "link_type": link_type,
        "status": "✅ link updated",
        "url": data.get("_links", {}).get("html", {}).get("href") or _work_item_edit_url(source_id),
    })


def ado_transition_work_item_state(inputs: dict) -> str:
    work_item_id = int(inputs["work_item_id"])
    target_state = str(inputs["state"]).strip()

    current = _get_work_item(work_item_id, expand="none")
    current_type = current.get("fields", {}).get("System.WorkItemType")
    if current_type in {"User Story", "Task"}:
        normalized_state = _normalize_user_story_task_state(target_state)
        if normalized_state is None:
            return json.dumps({
                "error": f"Invalid state for {current_type}. Allowed values: " + ", ".join(USER_STORY_TASK_STATES),
            })
        target_state = normalized_state

    patch = [{"op": "add", "path": "/fields/System.State", "value": target_state}]
    if inputs.get("reason"):
        patch.append({"op": "add", "path": "/fields/System.Reason", "value": inputs["reason"]})
    if inputs.get("comment"):
        patch.append({"op": "add", "path": "/fields/System.History", "value": inputs["comment"]})

    data = _update_work_item(work_item_id, patch)
    out = _summarize_item(data)
    out["status"] = "✅ state updated"
    return json.dumps(out)


def ado_tag_work_item(inputs: dict) -> str:
    work_item_id = int(inputs["work_item_id"])
    operation = inputs.get("operation", "add")
    normalize_map = inputs.get("normalize_map", {})

    current = _get_work_item(work_item_id, expand="none")
    fields = current.get("fields", {})
    existing_tags = [t.strip() for t in (fields.get("System.Tags") or "").split(";") if t.strip()]

    incoming = [t.strip() for t in inputs.get("tags", []) if t and t.strip()]

    if operation == "add":
        updated = sorted(set(existing_tags).union(incoming), key=str.lower)
    elif operation == "remove":
        remove_set = {t.lower() for t in incoming}
        updated = [t for t in existing_tags if t.lower() not in remove_set]
    elif operation == "replace":
        updated = sorted(set(incoming), key=str.lower)
    elif operation == "normalize":
        normalized = []
        for tag in existing_tags:
            normalized.append(normalize_map.get(tag, normalize_map.get(tag.lower(), tag)))
        updated = sorted({t.strip() for t in normalized if t and t.strip()}, key=str.lower)
    else:
        return json.dumps({"error": "Invalid operation. Use add, remove, replace, normalize."})

    patched = _update_work_item(work_item_id, [{
        "op": "add",
        "path": "/fields/System.Tags",
        "value": "; ".join(updated),
    }])
    out = _summarize_item(patched)
    out["operation"] = operation
    out["tags"] = updated
    out["status"] = "✅ tags updated"
    return json.dumps(out)


TOOL_HANDLERS = {
    "create_epic": ado_create_epic,
    "get_epic": ado_get_epic,
    "update_epic": ado_update_epic,
    "delete_epic": ado_delete_epic,
    "get_work_item_children": ado_get_work_item_children,
    "create_feature": ado_create_feature,
    "get_feature": ado_get_feature,
    "get_feature_children": ado_get_feature_children,
    "get_work_item_relations": ado_get_work_item_relations,
    "update_feature": ado_update_feature,
    "delete_feature": ado_delete_feature,
    "create_user_story": ado_create_user_story,
    "update_user_story": ado_update_user_story,
    "get_user_story":    ado_get_user_story,
    "delete_user_story": ado_delete_user_story,
    "add_comment":       ado_add_comment,
    "create_bug": ado_create_bug,
    "get_bug": ado_get_bug,
    "update_bug": ado_update_bug,
    "delete_bug": ado_delete_bug,
    "create_task": ado_create_task,
    "get_task": ado_get_task,
    "update_task": ado_update_task,
    "delete_task": ado_delete_task,
    "link_feature_hierarchy": ado_link_feature_hierarchy,
    "transition_work_item_state": ado_transition_work_item_state,
    "tag_work_item": ado_tag_work_item,
}

TOOLS = [
    {
        "name": "create_epic",
        "description": "Create a new Azure DevOps Epic.",
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "description": {"type": "string"},
                "acceptance_criteria": {"type": "string"},
                "priority": {"type": "integer", "description": "1=Critical, 2=High, 3=Medium, 4=Low"},
                "state": {"type": "string"},
                "area_path": {"type": "string"},
                "iteration_path": {"type": "string"},
                "tags": {"type": "string", "description": "Semicolon-separated"},
                "assignee": {"type": "string"},
                "start_date": {"type": "string", "description": "YYYY-MM-DD"},
                "target_date": {"type": "string", "description": "YYYY-MM-DD"},
                "business_value": {"type": "number"},
                "effort": {"type": "number"},
                "risk": {"type": "string", "description": "1 - High | 2 - Medium | 3 - Low"},
            },
            "required": ["title"],
        },
    },
    {
        "name": "get_epic",
        "description": "Retrieve an Epic by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "epic_id": {"type": "integer"},
            },
            "required": ["epic_id"],
        },
    },
    {
        "name": "update_epic",
        "description": "Update Epic fields by ID. Only send fields explicitly requested by the user.",
        "input_schema": {
            "type": "object",
            "properties": {
                "epic_id": {"type": "integer"},
                "title": {"type": "string"},
                "description": {"type": "string"},
                "acceptance_criteria": {"type": "string"},
                "priority": {"type": "integer", "description": "1=Critical, 2=High, 3=Medium, 4=Low"},
                "state": {"type": "string"},
                "area_path": {"type": "string"},
                "iteration_path": {"type": "string"},
                "tags": {"type": "string", "description": "Semicolon-separated"},
                "assignee": {"type": "string"},
                "start_date": {"type": "string", "description": "YYYY-MM-DD"},
                "target_date": {"type": "string", "description": "YYYY-MM-DD"},
                "business_value": {"type": "number"},
                "effort": {"type": "number"},
                "risk": {"type": "string", "description": "1 - High | 2 - Medium | 3 - Low"},
            },
            "required": ["epic_id"],
        },
    },
    {
        "name": "delete_epic",
        "description": "Delete, restore, or permanently delete an Epic by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "epic_id": {"type": "integer"},
                "action": {"type": "string", "description": "soft-delete | restore | permanent-delete"},
            },
            "required": ["epic_id"],
        },
    },
    {
        "name": "create_feature",
        "description": "Create a new Azure DevOps Feature with optional parent Epic link.",
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "parent_epic_id": {"type": "integer"},
                "description": {"type": "string"},
                "acceptance_criteria": {"type": "string"},
                "priority": {"type": "integer", "description": "1=Critical, 2=High, 3=Medium, 4=Low"},
                "state": {"type": "string"},
                "area_path": {"type": "string"},
                "iteration_path": {"type": "string"},
                "tags": {"type": "string", "description": "Semicolon-separated"},
                "assignee": {"type": "string"},
                "start_date": {"type": "string", "description": "YYYY-MM-DD"},
                "target_date": {"type": "string", "description": "YYYY-MM-DD"},
                "business_value": {"type": "number"},
                "effort": {"type": "number"},
                "risk": {"type": "string", "description": "1 - High | 2 - Medium | 3 - Low"},
            },
            "required": ["title"],
        },
    },
    {
        "name": "get_feature",
        "description": "Retrieve a Feature by ID. Response includes child_story_ids (list of child work item IDs) and child_story_count.",
        "input_schema": {
            "type": "object",
            "properties": {
                "feature_id": {"type": "integer"},
            },
            "required": ["feature_id"],
        },
    },
    {
        "name": "get_feature_children",
        "description": "Fetch full details of all child work items (User Stories, Bugs, etc.) linked under a Feature.",
        "input_schema": {
            "type": "object",
            "properties": {
                "feature_id": {"type": "integer", "description": "The Feature work item ID"},
            },
            "required": ["feature_id"],
        },
    },
    {
        "name": "get_work_item_children",
        "description": (
            "Fetch full report details for child work items under any Azure DevOps work item. "
            "Use this for Excel, CSV, TXT, or JSON reports. "
            "Use this for tasks under a story, bugs under a story, stories under a feature, "
            "features under an epic, or complete hierarchy reports. "
            "Use recursive=true for full hierarchy. "
            "Use item_types to filter, for example ['Task'], ['Bug'], ['User Story'], ['Feature']."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "work_item_id": {
                    "type": "integer",
                    "description": "The parent work item ID. Can be Epic, Feature, User Story, Bug, Task, or any work item."
                },
                "recursive": {
                    "type": "boolean",
                    "description": "True to fetch all descendants. False to fetch only direct children."
                },
                "include_parent": {
                    "type": "boolean",
                    "description": "True to include the parent work item as the first row."
                },
                "item_types": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Optional filter. Example: ['Task'], ['Bug'], ['User Story'], ['Feature'], ['Epic']."
                },
            },
            "required": ["work_item_id"],
        },
    },  
    {
        "name": "get_work_item_relations",
        "description": "Debug-only tool. Returns raw ADO relation types on a work item. Do not use for Excel, CSV, TXT, JSON, reports, task reports, child reports, or hierarchy reports.",
        "input_schema": {
            "type": "object",
            "properties": {
                "work_item_id": {"type": "integer"},
            },
            "required": ["work_item_id"],
        },
    },
    {
        "name": "update_feature",
        "description": "Update Feature fields by ID, including optional parent Epic re-linking.",
        "input_schema": {
            "type": "object",
            "properties": {
                "feature_id": {"type": "integer"},
                "title": {"type": "string"},
                "parent_epic_id": {"type": "integer"},
                "description": {"type": "string"},
                "acceptance_criteria": {"type": "string"},
                "priority": {"type": "integer", "description": "1=Critical, 2=High, 3=Medium, 4=Low"},
                "state": {"type": "string"},
                "area_path": {"type": "string"},
                "iteration_path": {"type": "string"},
                "tags": {"type": "string", "description": "Semicolon-separated"},
                "assignee": {"type": "string"},
                "start_date": {"type": "string", "description": "YYYY-MM-DD"},
                "target_date": {"type": "string", "description": "YYYY-MM-DD"},
                "business_value": {"type": "number"},
                "effort": {"type": "number"},
                "risk": {"type": "string", "description": "1 - High | 2 - Medium | 3 - Low"},
            },
            "required": ["feature_id"],
        },
    },
    {
        "name": "delete_feature",
        "description": "Delete, restore, or permanently delete a Feature by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "feature_id": {"type": "integer"},
                "action": {"type": "string", "description": "soft-delete | restore | permanent-delete"},
            },
            "required": ["feature_id"],
        },
    },
    {
        "name": "create_user_story",
        "description": "Create a new Azure DevOps user story using either a feature ID or a supported APAC project name.",
        "input_schema": {
            "type": "object",
            "properties": {
                "feature_id":          {"type": "integer", "description": "Parent feature work item ID (optional)"},
                "project_name":        {"type": "string", "description": "APAC project name for default area path"},
                "title":               {"type": "string"},
                "description":         {"type": "string"},
                "acceptance_criteria": {"type": "string"},
                "story_points":        {"type": "number"},
                "priority":            {"type": "integer", "description": "1=Critical 2=High 3=Medium 4=Low"},
                "assignee":            {"type": "string"},
                "tags":                {"type": "string", "description": "Semicolon-separated"},
                "area_path":           {"type": "string", "description": "Defaults to LexisNexis\\APAC\\{project_name}"},
                "iteration_path":      {"type": "string", "description": "Defaults to LexisNexis\\LexisNexis 3 Week M-F"},
            },
            "required": ["title"],
        },
    },
    {
        "name": "update_user_story",
        "description": "Update one or more fields of an existing Azure DevOps user story.",
        "input_schema": {
            "type": "object",
            "properties": {
                "story_id":            {"type": "integer"},
                "title":               {"type": "string"},
                "description":         {"type": "string"},
                "acceptance_criteria": {"type": "string"},
                "story_points":        {"type": "number"},
                "priority":            {"type": "integer", "description": "1=Critical, 2=High, 3=Medium, 4=Low"},
                "assignee":            {"type": "string"},
                "tags":                {"type": "string"},
                "state":               {"type": "string", "description": "For User Story use: Not Started / Ready / In Progress / Done / Closed / Removed"},
                "reason":              {"type": "string", "description": "Reason for the state. E.g. Approved, Implementation started, Merged, Accepted, Deferred, Removed from the backlog"},
                "area_path":           {"type": "string"},
                "iteration_path":      {"type": "string"},
                "comment":             {"type": "string", "description": "If provided, a comment is posted to the work item discussion after the field update."},
            },
            "required": ["story_id"],
        },
    },
    {
        "name": "get_user_story",
        "description": "Fetch all details of an Azure DevOps user story by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "story_id": {"type": "integer"},
            },
            "required": ["story_id"],
        },
    },
    {
        "name": "delete_user_story",
        "description": "Delete, restore, or permanently delete a user story by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "story_id": {"type": "integer"},
                "action": {"type": "string", "description": "soft-delete | restore | permanent-delete"},
            },
            "required": ["story_id"],
        },
    },
    {
        "name": "add_comment",
        "description": "Post a comment/discussion entry to any Azure DevOps work item (user story, bug, task, etc.).",
        "input_schema": {
            "type": "object",
            "properties": {
                "work_item_id": {"type": "integer", "description": "ID of the work item to comment on"},
                "text":         {"type": "string",  "description": "The comment text to post"},
            },
            "required": ["work_item_id", "text"],
        },
    },
    {
        "name": "create_bug",
        "description": "Create a new Azure DevOps bug.",
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "description": {"type": "string"},
                "repro_steps": {"type": "string"},
                "state": {"type": "string"},
                "priority": {"type": "integer"},
                "severity": {"type": "string"},
                "assignee": {"type": "string"},
                "tags": {"type": "string", "description": "Semicolon-separated"},
                "area_path": {"type": "string"},
                "iteration_path": {"type": "string"},
            },
            "required": ["title"],
        },
    },
    {
        "name": "get_bug",
        "description": "Retrieve a bug by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "bug_id": {"type": "integer"},
            },
            "required": ["bug_id"],
        },
    },
    {
        "name": "update_bug",
        "description": "Update bug fields by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "bug_id": {"type": "integer"},
                "title": {"type": "string"},
                "description": {"type": "string"},
                "repro_steps": {"type": "string"},
                "priority": {"type": "integer"},
                "severity": {"type": "string"},
                "assignee": {"type": "string"},
                "tags": {"type": "string"},
                "state": {"type": "string"},
                "area_path": {"type": "string"},
                "iteration_path": {"type": "string"},
            },
            "required": ["bug_id"],
        },
    },
    {
        "name": "delete_bug",
        "description": "Delete, restore, or permanently delete a bug by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "bug_id": {"type": "integer"},
                "action": {"type": "string", "description": "soft-delete | restore | permanent-delete"},
            },
            "required": ["bug_id"],
        },
    },
    {
        "name": "create_task",
        "description": "Create a task with optional parent link. If parent_id is a User Story, Area Path and Iteration Path inherit from parent unless explicitly provided.",
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "description": {"type": "string"},
                "priority": {"type": "integer"},
                "assignee": {"type": "string"},
                "tags": {"type": "string"},
                "remaining_work": {"type": "number"},
                "parent_id": {"type": "integer"},
                "area_path": {"type": "string", "description": "Optional override. If omitted and parent is User Story, inherits parent Area Path."},
                "iteration_path": {"type": "string", "description": "Optional override. If omitted and parent is User Story, inherits parent Iteration Path."},
            },
            "required": ["title"],
        },
    },
    {
        "name": "get_task",
        "description": "Retrieve a task by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "task_id": {"type": "integer"},
            },
            "required": ["task_id"],
        },
    },
    {
        "name": "update_task",
        "description": "Update task fields by ID. Supports all standard task fields including work estimates and comments.",
        "input_schema": {
            "type": "object",
            "properties": {
                "task_id":           {"type": "integer"},
                "title":             {"type": "string"},
                "description":       {"type": "string"},
                "state":             {"type": "string", "description": "Not Started / Ready / In Progress / Done / Closed / Removed"},
                "reason":            {"type": "string", "description": "Reason tied to current state, e.g. Approved, Implementation started, Accepted"},
                "priority":          {"type": "integer", "description": "1=Critical, 2=High, 3=Medium, 4=Low"},
                "activity":          {"type": "string", "description": "Work type: Deployment, Design, Development, Documentation, Requirements, Testing"},
                "assignee":          {"type": "string"},
                "area_path":         {"type": "string"},
                "iteration_path":    {"type": "string"},
                "original_estimate": {"type": "number", "description": "Planned hours"},
                "remaining_work":    {"type": "number", "description": "Hours of work remaining"},
                "completed_work":    {"type": "number", "description": "Hours already completed"},
                "tags":              {"type": "string", "description": "Semicolon-separated tags"},
                "comment":           {"type": "string", "description": "Post a discussion comment alongside any field updates"},
            },
            "required": ["task_id"],
        },
    },
    {
        "name": "delete_task",
        "description": "Delete, restore, or permanently delete a task by ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "task_id": {"type": "integer"},
                "action": {"type": "string", "description": "soft-delete | restore | permanent-delete"},
            },
            "required": ["task_id"],
        },
    },
    {
        "name": "link_feature_hierarchy",
        "description": "Link or unlink work items using sourceId/targetId and linkType (parent, dependency, related).",
        "input_schema": {
            "type": "object",
            "properties": {
                "action": {"type": "string", "description": "link | unlink"},
                "sourceId": {"type": "integer", "description": "Item you are updating"},
                "targetId": {"type": "integer", "description": "Item you are linking to/from"},
                "linkType": {"type": "string", "description": "parent | dependency | related"},
                "comment": {"type": "string", "description": "Optional link comment; used for action=link"},
            },
            "required": ["action", "sourceId", "targetId", "linkType"],
        },
    },
    {
        "name": "transition_work_item_state",
        "description": "Transition a work item to a new state.",
        "input_schema": {
            "type": "object",
            "properties": {
                "work_item_id": {"type": "integer"},
                "state": {"type": "string", "description": "For User Story/Task use: Not Started / Ready / In Progress / Done / Closed / Removed"},
                "reason": {"type": "string"},
                "comment": {"type": "string"},
            },
            "required": ["work_item_id", "state"],
        },
    },
    {
        "name": "tag_work_item",
        "description": "Add, remove, replace, or normalize tags on a work item.",
        "input_schema": {
            "type": "object",
            "properties": {
                "work_item_id": {"type": "integer"},
                "operation": {"type": "string", "description": "add | remove | replace | normalize"},
                "tags": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Tag list for add/remove/replace",
                },
                "normalize_map": {
                    "type": "object",
                    "description": "Mapping of old tag to canonical tag for normalize",
                },
            },
            "required": ["work_item_id", "operation"],
        },
    },
]
SYSTEM_PROMPT = """You are an Azure DevOps assistant. Help users manage Epics, Features, User Stories, Bugs, and Tasks with CRUD-style operations, hierarchy links, state transitions, comments, and tags.

Critical backend execution rules:
- For any create, update, delete, restore, comment, link, unlink, tag, or state-transition request, you MUST call the matching Azure DevOps tool.
- Never say an operation succeeded unless you received a successful tool_result from the matching tool.
- If a required field is missing, ask for the missing field instead of pretending the operation succeeded.
- Do not infer that an operation was completed from previous chat messages.
- Do not produce a success confirmation from memory.
- After a tool call, summarize only what the tool_result confirms.
- For hierarchy output in chat, use clean markdown bullet lists or markdown tables. Do not use HTML entities like &nbsp;.

General rules:
- For create operations: collect required IDs/titles and ask for missing fields.
- For update operations: only send fields the user explicitly mentioned. Never guess values.
- For delete operations: default to soft-delete unless user explicitly asks for permanent-delete.
- For restore operations: use the same item ID with action restore.
- For create_user_story: require at least one of feature_id or project_name.
- For create_user_story: if project_name is present, it must be one of: UK BES, Irish Project, CA Publisher, Team Aryan India, ACP-APAC Forms and Precedents, Bold, Content Build Asia and Pacific, MNCR - Migration.
- For create_user_story defaults unless user overrides: story_points=3, priority=3, iteration_path=LexisNexis\\LexisNexis 3 Week M-F, area_path=LexisNexis\\APAC\\{project_name}.
- For create_task: when parent_id points to a User Story, inherit Area Path and Iteration Path from parent unless user explicitly overrides.
- For User Story and Task state updates, use only: Not Started, Ready, In Progress, Done, Closed, Removed.
- To list child work items for user-facing reports, prefer get_work_item_children. Use get_feature_children only for simple feature child lookup.
- get_feature returns child_story_ids IDs only; get_feature_children returns full details of each child.
- Always confirm results with the work item ID, title if available, state if available, and URL if available.
- Format responses clearly using markdown.
- Use ✅ only when a backend tool_result confirms success.
- Use ❌ when a backend tool_result contains an error.
- For Excel, CSV, TXT, or JSON reports of child work items, use get_work_item_children.
- To create reports for child work items under any Epic, Feature, User Story, Bug, Task, or work item, use get_work_item_children.
- For complete hierarchy reports, use get_work_item_children with recursive=true and include_parent=true.
- For direct child reports, use get_work_item_children with recursive=false.
- For task-only reports, use get_work_item_children with item_types=["Task"].
- For bug-only reports, use get_work_item_children with item_types=["Bug"].
- Do not use get_work_item_relations for user-facing Excel, CSV, TXT, or JSON reports unless the user explicitly asks for raw relations.
"""

# ── Agentic loop ──────────────────────────────────────────────────────────────

class ToolEvent(TypedDict, total=False):
    tool: str
    input: dict
    success: bool
    mutating: bool
    result: dict
    raw_result: str
    refresh_ids: list[int]
    verified_work_items: list[dict]
    error: str


class AgentState(TypedDict, total=False):
    messages: Annotated[list[dict], add]
    tool_events: Annotated[list[ToolEvent], add]


MUTATING_TOOLS = {
    "create_epic",
    "update_epic",
    "delete_epic",
    "create_feature",
    "update_feature",
    "delete_feature",
    "create_user_story",
    "update_user_story",
    "delete_user_story",
    "create_bug",
    "update_bug",
    "delete_bug",
    "create_task",
    "update_task",
    "delete_task",
    "add_comment",
    "link_feature_hierarchy",
    "transition_work_item_state",
    "tag_work_item",
}


DELETE_TOOLS = {
    "delete_epic",
    "delete_feature",
    "delete_user_story",
    "delete_bug",
    "delete_task",
}


def _extract_final_text(messages: list[dict]) -> str:
    for message in reversed(messages):
        if message.get("role") != "assistant":
            continue

        content = message.get("content")

        if isinstance(content, str):
            text = content.strip()
            if text:
                return text

        if isinstance(content, list):
            text_blocks = [
                b.get("text", "")
                for b in content
                if isinstance(b, dict) and b.get("type") == "text"
            ]
            text = "\n".join(t for t in text_blocks if t).strip()
            if text:
                return text

    return "Done."


def _safe_json_loads(value: str):
    try:
        return json.loads(value)
    except Exception:
        return {"raw": value}


def _content_to_text(content) -> str:
    if isinstance(content, str):
        return content

    if isinstance(content, list):
        parts = []
        for block in content:
            if isinstance(block, dict):
                if block.get("type") == "text":
                    parts.append(str(block.get("text", "")))
                elif isinstance(block.get("content"), str):
                    parts.append(block["content"])
            elif isinstance(block, str):
                parts.append(block)
        return "\n".join(p for p in parts if p).strip()

    return ""


def _sanitize_client_messages(messages: list[dict]) -> list[dict]:
    """
    Frontend should not send old tool_use/tool_result blocks back to the model.
    This function also trims history to avoid old success messages causing false confirmations.
    """
    clean: list[dict] = []

    for msg in messages:
        role = msg.get("role")
        if role not in {"user", "assistant"}:
            continue

        content = msg.get("content", "")

        if isinstance(content, list):
            text_blocks = []
            for block in content:
                if not isinstance(block, dict):
                    continue

                # Drop tool internals from client-provided history.
                if block.get("type") in {"tool_use", "tool_result"}:
                    continue

                if block.get("type") == "text":
                    text_blocks.append(block.get("text", ""))

            content = "\n".join(t for t in text_blocks if t).strip()

        if not isinstance(content, str):
            content = str(content)

        content = content.strip()
        if not content:
            continue

        clean.append({"role": role, "content": content})

    return clean[-CHAT_HISTORY_WINDOW:]


def _latest_user_text(messages: list[dict]) -> str:
    for msg in reversed(messages):
        if msg.get("role") == "user":
            return _content_to_text(msg.get("content", ""))
    return ""


def _looks_like_mutation_request(text: str) -> bool:
    text = text.lower()

    mutation_words = (
        "create",
        "add",
        "update",
        "change",
        "edit",
        "set",
        "assign",
        "delete",
        "remove",
        "restore",
        "permanent-delete",
        "permanently delete",
        "comment",
        "link",
        "unlink",
        "tag",
        "transition",
        "move to",
        "mark as",
        "close",
        "reopen",
    )

    ado_words = (
        "story",
        "user story",
        "task",
        "bug",
        "feature",
        "epic",
        "work item",
        "ado",
        "azure devops",
        "description",
        "acceptance criteria",
        "state",
        "priority",
        "assignee",
        "tags",
    )

    return any(w in text for w in mutation_words) and any(w in text for w in ado_words)


def _looks_like_success_text(text: str) -> bool:
    lowered = text.lower()

    success_terms = (
        "✅",
        "created",
        "updated",
        "deleted",
        "restored",
        "comment posted",
        "linked",
        "unlinked",
        "state updated",
        "tags updated",
        "success",
        "completed",
        "done",
    )

    return any(term in lowered for term in success_terms)


def _is_success_result(parsed_result) -> bool:
    if isinstance(parsed_result, dict) and parsed_result.get("error"):
        return False
    return True


def _coerce_int(value) -> int | None:
    if value is None:
        return None

    try:
        return int(value)
    except Exception:
        return None


def _unique_ints(values: list[int | None]) -> list[int]:
    seen = set()
    out = []

    for value in values:
        if value is None:
            continue

        value = int(value)
        if value not in seen:
            seen.add(value)
            out.append(value)

    return out


def _extract_refresh_ids(tool_name: str, tool_input: dict, parsed_result: dict) -> list[int]:
    """
    IDs the UI may need to refresh after a backend mutation.
    """
    ids: list[int | None] = []

    if isinstance(parsed_result, dict):
        ids.append(_coerce_int(parsed_result.get("id")))
        ids.append(_coerce_int(parsed_result.get("work_item_id")))

    for key in (
        "work_item_id",
        "story_id",
        "user_story_id",
        "userStoryId",
        "task_id",
        "bug_id",
        "feature_id",
        "epic_id",
        "id",
    ):
        ids.append(_coerce_int(tool_input.get(key)))

    if tool_name == "link_feature_hierarchy":
        ids.append(_coerce_int(tool_input.get("sourceId") or tool_input.get("source_id") or tool_input.get("child_id")))
        ids.append(_coerce_int(tool_input.get("targetId") or tool_input.get("target_id") or tool_input.get("parent_id")))

    if tool_name == "create_user_story":
        ids.append(_coerce_int(tool_input.get("feature_id") or tool_input.get("parent_feature_id")))

    if tool_name == "create_task":
        ids.append(_coerce_int(tool_input.get("parent_id") or tool_input.get("story_id") or tool_input.get("user_story_id")))

    if tool_name == "create_feature":
        ids.append(_coerce_int(tool_input.get("parent_epic_id")))

    return _unique_ints(ids)


def _should_verify_after_tool(tool_name: str, tool_input: dict) -> bool:
    """
    Do not fetch after normal soft/permanent delete because the item may no longer be available.
    Restore should be verified.
    """
    if tool_name not in MUTATING_TOOLS:
        return False

    if tool_name in DELETE_TOOLS:
        action = str(tool_input.get("action", "soft-delete")).strip().lower()
        return action == "restore"

    return True


def _verified_work_item_snapshot(work_item_id: int) -> dict:
    data = _get_work_item(work_item_id, expand="all")
    fields = data.get("fields", {})

    snapshot = _summarize_item(data)
    snapshot.update({
        "description": fields.get("System.Description"),
        "acceptance_criteria": fields.get("Microsoft.VSTS.Common.AcceptanceCriteria"),
        "repro_steps": fields.get("Microsoft.VSTS.TCM.ReproSteps"),
        "reason": fields.get("System.Reason"),
        "area_path": fields.get("System.AreaPath"),
        "iteration_path": fields.get("System.IterationPath"),
        "story_points": fields.get("Microsoft.VSTS.Scheduling.StoryPoints"),
        "remaining_work": fields.get("Microsoft.VSTS.Scheduling.RemainingWork"),
        "original_estimate": fields.get("Microsoft.VSTS.Scheduling.OriginalEstimate"),
        "completed_work": fields.get("Microsoft.VSTS.Scheduling.CompletedWork"),
        "severity": fields.get("Microsoft.VSTS.Common.Severity"),
        "activity": fields.get("Microsoft.VSTS.Common.Activity"),
    })

    return snapshot


def _verified_snapshots_for_ids(ids: list[int]) -> list[dict]:
    verified = []

    for work_item_id in ids:
        try:
            verified.append(_verified_work_item_snapshot(work_item_id))
        except requests.HTTPError as e:
            verified.append({
                "id": work_item_id,
                "verification_error": f"{e.response.status_code} {e.response.text}",
            })
        except Exception as e:
            verified.append({
                "id": work_item_id,
                "verification_error": str(e),
            })

    return verified


def _llm_node(state: AgentState) -> AgentState:
    response = client.messages.create(
        model=LLM_MODEL,
        max_tokens=MAX_TOKENS,
        temperature=TEMPERATURE,
        system=SYSTEM_PROMPT,
        tools=TOOLS,
        messages=state["messages"],
    )

    assistant_content = []

    for block in response.content:
        if block.type == "text":
            assistant_content.append({
                "type": "text",
                "text": block.text,
            })
        elif block.type == "tool_use":
            assistant_content.append({
                "type": "tool_use",
                "id": block.id,
                "name": block.name,
                "input": dict(block.input),
            })

    return {
        "messages": [
            {
                "role": "assistant",
                "content": assistant_content,
            }
        ]
    }


def _has_tool_calls(state: AgentState) -> str:
    if not state.get("messages"):
        return "done"

    last_message = state["messages"][-1]
    if last_message.get("role") != "assistant":
        return "done"

    content = last_message.get("content")
    if not isinstance(content, list):
        return "done"

    for block in content:
        if isinstance(block, dict) and block.get("type") == "tool_use":
            return "tools"

    return "done"


def _tools_node(state: AgentState) -> AgentState:
    last_message = state["messages"][-1]
    content = last_message.get("content", [])

    tool_results = []
    tool_events: list[ToolEvent] = []

    for block in content:
        if not isinstance(block, dict) or block.get("type") != "tool_use":
            continue

        tool_name = block.get("name")
        tool_input = block.get("input", {}) or {}
        tool_use_id = block.get("id")

        handler = TOOL_HANDLERS.get(tool_name)

        if handler is None:
            raw_result = json.dumps({"error": f"Unknown tool: {tool_name}"})
        else:
            try:
                raw_result = handler(dict(tool_input))
            except requests.HTTPError as e:
                raw_result = json.dumps({
                    "error": f"ADO API error: {e.response.status_code} {e.response.text}",
                })
            except Exception as e:
                raw_result = json.dumps({"error": str(e)})

        parsed_result = _safe_json_loads(raw_result)
        success = _is_success_result(parsed_result)
        mutating = tool_name in MUTATING_TOOLS

        refresh_ids: list[int] = []
        verified_work_items: list[dict] = []

        if isinstance(parsed_result, dict):
            refresh_ids = _extract_refresh_ids(tool_name, tool_input, parsed_result)

        if success and mutating and _should_verify_after_tool(tool_name, tool_input):
            verified_work_items = _verified_snapshots_for_ids(refresh_ids)

            if isinstance(parsed_result, dict):
                parsed_result = dict(parsed_result)
                parsed_result["verified_work_items"] = verified_work_items
                raw_result = json.dumps(parsed_result)

        event: ToolEvent = {
            "tool": str(tool_name),
            "input": dict(tool_input),
            "success": success,
            "mutating": mutating,
            "result": parsed_result if isinstance(parsed_result, dict) else {"raw": parsed_result},
            "raw_result": raw_result,
            "refresh_ids": refresh_ids,
            "verified_work_items": verified_work_items,
        }

        if isinstance(parsed_result, dict) and parsed_result.get("error"):
            event["error"] = str(parsed_result.get("error"))

        tool_events.append(event)

        tool_results.append({
            "type": "tool_result",
            "tool_use_id": tool_use_id,
            "content": raw_result,
        })

    return {
        "messages": [
            {
                "role": "user",
                "content": tool_results,
            }
        ],
        "tool_events": tool_events,
    }


def _build_agent_graph():
    graph = StateGraph(AgentState)

    graph.add_node("llm", _llm_node)
    graph.add_node("tools", _tools_node)

    graph.set_entry_point("llm")

    graph.add_conditional_edges(
        "llm",
        _has_tool_calls,
        {
            "tools": "tools",
            "done": END,
        },
    )

    graph.add_edge("tools", "llm")

    return graph.compile()


AGENT_GRAPH = _build_agent_graph()


def _invoke_graph(messages: list[dict]) -> AgentState:
    return AGENT_GRAPH.invoke(
        {
            "messages": list(messages),
            "tool_events": [],
        },
        config={"recursion_limit": 40},
    )


def _collect_run_metadata(final_state: AgentState) -> dict:
    events = final_state.get("tool_events", []) or []

    tool_executed = len(events) > 0
    mutation_executed = any(
        event.get("mutating") and event.get("success")
        for event in events
    )

    failed_tools = [
        {
            "tool": event.get("tool"),
            "error": event.get("error"),
            "input": event.get("input"),
        }
        for event in events
        if not event.get("success")
    ]

    updated_ids = sorted({
        int(work_item_id)
        for event in events
        if event.get("mutating") and event.get("success")
        for work_item_id in event.get("refresh_ids", [])
        if work_item_id is not None
    })

    verified_work_items_by_id: dict[int, dict] = {}

    for event in events:
        for item in event.get("verified_work_items", []) or []:
            item_id = item.get("id")
            if item_id is not None:
                verified_work_items_by_id[int(item_id)] = item

    structured_outputs = [
        {
            "tool": event.get("tool"),
            "success": event.get("success"),
            "mutating": event.get("mutating"),
            "result": event.get("result"),
        }
        for event in events
    ]
    
    return {
        "tool_executed": tool_executed,
        "mutation_executed": mutation_executed,
        "structured_outputs": structured_outputs,
        "tool_calls": [
            {
                "tool": event.get("tool"),
                "success": event.get("success"),
                "mutating": event.get("mutating"),
                "refresh_ids": event.get("refresh_ids", []),
                "error": event.get("error"),
            }
            for event in events
        ],
        "updated_work_item_ids": updated_ids,
        "verified_work_items": list(verified_work_items_by_id.values()),
        "requires_refresh": mutation_executed and bool(updated_ids),
        "failed_tools": failed_tools,
    }


def _false_success_guard(reply: str, latest_user_text: str, metadata: dict) -> str:
    """
    Prevent UI from showing fake success when no backend mutation happened.
    """
    if not _looks_like_mutation_request(latest_user_text):
        return reply

    if metadata.get("mutation_executed"):
        return reply

    if not _looks_like_success_text(reply):
        return reply

    return (
        "❌ No backend update was performed.\n\n"
        "The assistant generated a success-style response, but no successful Azure DevOps mutation tool call was recorded. "
        "I cannot confirm the change was saved to ADO. Please resend the request with the work item ID and exact field change."
    )


def run_agent(messages: list[dict]) -> dict:
    clean_messages = _sanitize_client_messages(messages)
    latest_user = _latest_user_text(clean_messages)

    final_state = _invoke_graph(clean_messages)
    reply = _extract_final_text(final_state.get("messages", []))
    metadata = _collect_run_metadata(final_state)

    # Repair pass:
    # If the model tried to claim success without a mutation tool, force one retry.
    if (
        _looks_like_mutation_request(latest_user)
        and not metadata.get("mutation_executed")
        and _looks_like_success_text(reply)
    ):
        repair_messages = clean_messages + [
            {
                "role": "user",
                "content": (
                    "Internal correction: Your previous response appeared to confirm a backend change, "
                    "but no Azure DevOps mutation tool was called. For my original request, call the correct ADO tool now "
                    "if all required fields are present. If required fields are missing, ask me for them. "
                    "Do not claim success without a successful tool_result."
                ),
            }
        ]

        final_state = _invoke_graph(repair_messages)
        reply = _extract_final_text(final_state.get("messages", []))
        metadata = _collect_run_metadata(final_state)

    reply = _false_success_guard(reply, latest_user, metadata)

    return {
        "reply": reply,
        **metadata,
    }
    
#  export helper functions  
EXPORT_PREFERRED_COLUMNS = [
    "level",
    "parent_id",
    "id",
    "type",
    "title",
    "state",
    "reason",
    "priority",
    "assignee",
    "story_points",
    "activity",
    "original_estimate",
    "completed_work",
    "remaining_work",
    "severity",
    "business_value",
    "effort",
    "risk",
    "start_date",
    "target_date",
    "tags",
    "area_path",
    "iteration_path",
    "url",
    "status",
    "tool",
    "success",
    "mutating",
    "error",
]

def _detect_requested_output_format(text: str) -> str | None:
    """
    Export only when user explicitly asks for a specific output format.
    """
    if not text:
        return None

    normalized = f" {text.lower()} "

    has_export_intent = any(
        word in normalized
        for word in (
            " export ",
            " download ",
            " file ",
            " format ",
            " as ",
            " in ",
            " generate ",
            " give me ",
            " provide ",
        )
    )

    if not has_export_intent:
        return None

    if any(word in normalized for word in (" excel ", " xlsx ", " .xlsx ")):
        return "xlsx"

    if any(word in normalized for word in (" csv ", " .csv ")):
        return "csv"

    if any(word in normalized for word in (" txt ", " text file ", " .txt ")):
        return "txt"

    if any(word in normalized for word in (" json ", " .json ")):
        return "json"

    return None


def _file_name(extension: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"ado_output_{stamp}.{extension}"


def _cell_value(value):
    """
    Convert nested values into safe CSV/Excel/TXT cell values.
    """
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)

    return value


def _collect_columns(rows: list[dict]) -> list[str]:
    all_keys = []

    for row in rows:
        for key in row.keys():
            if key not in all_keys:
                all_keys.append(key)

    preferred = [col for col in EXPORT_PREFERRED_COLUMNS if col in all_keys]
    remaining = [col for col in all_keys if col not in preferred]

    return preferred + remaining


def _normalize_rows(value) -> list[dict]:
    """
    Convert dict/list/scalar data into list[dict] for CSV and Excel.
    """
    if not value:
        return []

    if isinstance(value, list):
        rows = []

        for item in value:
            if isinstance(item, dict):
                rows.append(item)
            else:
                rows.append({"value": item})

        return rows

    if isinstance(value, dict):
        return [value]

    return [{"value": value}]


def _pick_export_rows(result: dict) -> tuple[list[dict], str]:
    """
    Select best structured data from run_agent result.

    Priority:
    1. verified_work_items
    2. tool result arrays like rows, children, items
    3. latest single tool result
    4. tool_calls
    5. reply fallback
    """

    verified = result.get("verified_work_items") or []
    if verified:
        return _normalize_rows(verified), "verified_work_items"

    structured_outputs = result.get("structured_outputs") or []

    for output in reversed(structured_outputs):
        tool_result = output.get("result")

        if not isinstance(tool_result, dict):
            continue

        for key in (
            "rows",
            "children",
            "items",
            "work_items",
            "verified_work_items",
            "failed_tools",
            "tool_calls",
        ):
            value = tool_result.get(key)

            if isinstance(value, list) and value:
                return _normalize_rows(value), key

    for output in reversed(structured_outputs):
        tool_result = output.get("result")

        if isinstance(tool_result, dict):
            row = {
                "tool": output.get("tool"),
                **tool_result,
            }
            return _normalize_rows(row), output.get("tool") or "tool_result"

    tool_calls = result.get("tool_calls") or []
    if tool_calls:
        return _normalize_rows(tool_calls), "tool_calls"

    return [{"reply": result.get("reply", "")}], "reply"


def _export_txt(result: dict):
    rows, dataset_name = _pick_export_rows(result)

    lines = []
    lines.append("ADO Agent Output")
    lines.append(f"Dataset: {dataset_name}")
    lines.append("")

    if result.get("reply"):
        lines.append("Reply:")
        lines.append(str(result["reply"]))
        lines.append("")

    lines.append("Structured Data:")
    lines.append(json.dumps(rows, indent=2, ensure_ascii=False))

    content = "\n".join(lines)

    return StreamingResponse(
        iter([content.encode("utf-8")]),
        media_type="text/plain",
        headers={
            "Content-Disposition": f'attachment; filename="{_file_name("txt")}"'
        },
    )


def _export_csv(result: dict):
    rows, dataset_name = _pick_export_rows(result)

    if not rows:
        rows = [{"message": "No structured data available"}]

    columns = _collect_columns(rows)

    output = io.StringIO()

    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()

    for row in rows:
        writer.writerow({
            col: _cell_value(row.get(col))
            for col in columns
        })

    content = output.getvalue()

    return StreamingResponse(
        iter([content.encode("utf-8")]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{_file_name("csv")}"'
        },
    )


def _export_xlsx(result: dict):
    rows, dataset_name = _pick_export_rows(result)

    if not rows:
        rows = [{"message": "No structured data available"}]

    columns = _collect_columns(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = dataset_name[:31] or "Output"

    sheet.append(columns)

    for row in rows:
        sheet.append([
            _cell_value(row.get(col))
            for col in columns
        ])

    for col_idx, column_name in enumerate(columns, start=1):
        max_length = len(str(column_name))

        for row_idx in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value

            if value is not None:
                max_length = max(max_length, len(str(value)))

        column_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{_file_name("xlsx")}"'
        },
    )
    
# ── Routes ────────────────────────────────────────────────────────────────────

class ChatRequest(BaseModel):
    messages: list[dict]
    output_format: str | None = None


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/debug/relations/{work_item_id}")
def debug_relations(work_item_id: int):
    """Return the raw relations array from ADO for a work item — bypasses the agent."""
    try:
        data = _get_work_item(work_item_id)
        relations = data.get("relations") or []
        return {
            "work_item_id": work_item_id,
            "relation_count": len(relations),
            "relations": [
                {"rel": r.get("rel"), "url": r.get("url"), "attributes": r.get("attributes")}
                for r in relations
            ],
        }
    except requests.HTTPError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)

@app.post("/chat")
def chat(req: ChatRequest):
    try:
        result = run_agent(req.messages)

        latest_user_text = _latest_user_text(req.messages)

        requested_format = req.output_format

        if requested_format:
            requested_format = requested_format.strip().lower()
        else:
            requested_format = _detect_requested_output_format(latest_user_text)

        # Default behavior: normal chat response
        if requested_format is None:
            return result

        # Explicit JSON request still returns API JSON
        if requested_format == "json":
            return result

        if requested_format in {"txt", "text"}:
            return _export_txt(result)

        if requested_format == "csv":
            return _export_csv(result)

        if requested_format in {"xlsx", "excel"}:
            return _export_xlsx(result)

        raise HTTPException(
            status_code=400,
            detail="Unsupported output format. Use json, txt, csv, or xlsx.",
        )

    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=str(e))
